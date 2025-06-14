from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_job_tracker():
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create sheets
    applications = wb.create_sheet("Applications Overview")
    interviews = wb.create_sheet("Interviews")
    follow_ups = wb.create_sheet("Follow-ups")
    companies = wb.create_sheet("Companies Research")
    skills = wb.create_sheet("Skills & Requirements")
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Applications Overview Sheet
    app_headers = [
        "Company Name", "Position", "Location", "Application Date", 
        "Application Status", "Application Method", "Job Posting URL", "Notes"
    ]
    for col, header in enumerate(app_headers, 1):
        cell = applications.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        applications.column_dimensions[get_column_letter(col)].width = 20
    
    # Interviews Sheet
    interview_headers = [
        "Company Name", "Position", "Interview Date", "Interview Type",
        "Interviewers", "Key Questions Asked", "Your Performance Notes", "Follow-up Required"
    ]
    for col, header in enumerate(interview_headers, 1):
        cell = interviews.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        interviews.column_dimensions[get_column_letter(col)].width = 20
    
    # Follow-ups Sheet
    followup_headers = [
        "Company Name", "Contact Person", "Follow-up Date", "Method",
        "Response Received", "Next Steps"
    ]
    for col, header in enumerate(followup_headers, 1):
        cell = follow_ups.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        follow_ups.column_dimensions[get_column_letter(col)].width = 20
    
    # Companies Research Sheet
    company_headers = [
        "Company Name", "Industry", "Company Size", "Key Products/Services",
        "Notable Projects", "Company Culture Notes", "Glassdoor Rating", "LinkedIn Insights"
    ]
    for col, header in enumerate(company_headers, 1):
        cell = companies.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        companies.column_dimensions[get_column_letter(col)].width = 20
    
    # Skills & Requirements Sheet
    skills_headers = [
        "Position Type", "Required Skills", "Preferred Skills",
        "Your Match Level", "Skills to Develop"
    ]
    for col, header in enumerate(skills_headers, 1):
        cell = skills.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        skills.column_dimensions[get_column_letter(col)].width = 25
    
    # Add data validation for status
    status_list = ["Applied", "Interviewing", "Offer", "Rejected", "Withdrawn"]
    for row in range(2, 100):  # Add validation for first 100 rows
        cell = applications.cell(row=row, column=5)  # Status column
        cell.value = f'=IF(ISBLANK(A{row}),"",CHOOSE(MATCH(D{row},{"|".join(status_list)},0),{",".join(status_list)}))'
    
    # Add some example data
    example_data = [
        ["HSBC", "Quantitative Analyst", "Toronto", datetime.now().strftime("%Y-%m-%d"),
         "Applied", "LinkedIn", "https://example.com", "Entry-level quant role"],
        ["KPMG", "Financial Engineer", "Toronto", datetime.now().strftime("%Y-%m-%d"),
         "Interviewing", "Company Website", "https://example.com", "Valuations team"]
    ]
    
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            cell = applications.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left')
    
    # Save the workbook
    wb.save("public/templates/Job_Application_Tracker.xlsx")

if __name__ == "__main__":
    create_job_tracker() 